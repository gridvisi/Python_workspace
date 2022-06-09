#    原文链接：https: // blog.csdn.net / qq_39314932 / java / article / details / 96180914
# https://segmentfault.com/a/1190000005144821
from openpyxl import Workbook
wb = Workbook()
ws = wb.active             #默认创建第一个表，默认名字为sheet
ws1 = wb.create_sheet()    #创建第二个表
ws1.title = "New Title"    #为第二个表设置名字
ws2 = wb.get_sheet_by_name("New Title")                #通过名字获取表，和第二个表示一个表
ws1.save('your_name.xlsx') #保存

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
wb = Workbook()
dest_filename = 'empty_book.xlsx'
ws1 = wb.active              #第一个表
ws1.title = "range names"    #第一个表命名
#遍历第一个表的1到39行，每行赋值从1到599.
for row in range(1,40):
    ws1.append(range(600))
ws2 = wb.create_sheet(title="Pi") # 创建第二个表
ws2['F5'] = 3.14     #为第二个表的F5单元格赋值为3.14
ws3 = wb.create_sheet(title="Data")  #创建第三个表
#下面遍历第三个表的10到19行，27到53列，并对每一行的单元格赋一个当前列名的名字如下图

for row in range(10,20):
    for col in range(27,54):
        _=ws3.cell(column=col,row=row,value="%s" % get_column_letter(col)) #_当作一个普通的变量，一般表示后边不再使用
wb.save(filename=dest_filename) #保存

import time
tis1 =time.perf_counter()
print("等待2秒......")
time.sleep(2)
tis2=time.perf_counter()
import xlrd
from xlrd import xldate_as_tuple
import datetime
from openpyxl import Workbook
excel = xlrd.open_workbook("test.xls")
excel.sheet_names() # 获取excel里的工作表sheet名称数组
sheet = excel.sheet_by_index(0) #根据下标获取对应的sheet表

excel_path = "D:\\python\\excel\\tesla price.xls"

# 打开文件，获取excel文件的workbook（工作簿）对象
excel = xlrd.open_workbook(excel_path, encoding_override="utf-8")

# 获取sheet对象
all_sheet = excel.sheets()


# 循环遍历每个sheet对象
for sheet in all_sheet:
    print("该Excel共有{0}个sheet,当前sheet名称为{1},该sheet共有{2}行,{3}列"
    .format(len(all_sheet), sheet.name, sheet.nrows, sheet.ncols))
    sheet_cell_first_row = sheet.row(0)  # 获取指定行对象
    print(sheet_cell_first_row)
    for i in range(sheet.nrows):  # 依次遍历获得每一行对象
        each_cell_value_row = sheet.row(i)
    print(each_cell_value_row, type(each_cell_value_row))


