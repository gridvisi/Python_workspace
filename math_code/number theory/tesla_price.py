import openpyxl
import pandas as pd

import xlrd
from datetime import date
#coding=utf-8

import xlrd
# 打开 xls 文件道
book = xlrd.open_workbook(r'D:/python/excel/tesla_price.xls')
print("表单回数量:", book.nsheets)
print("表单名称:", book.sheet_names())
# 获取第 1 个表单
sh = book.sheet_by_index(0)
#print(u"表单 %s 共 %d 行答 %d 列" % (sh.name, sh.nrows, sh.ncols))
#print("第 11 行第 3 列:", sh.cell_value(10, 2))

prices = []
days = 21 #股价波动的时间跨度
for i in range(days,0,-1):
    prices.append(sh.cell_value(i, 1))
#print('tesla:',prices)

prices = [716.94, 742.0, 698.97, 590.16, 562.09, 554.2, 545.0, 511.2, 509.5, 481.03, 504.0, 501.25, 510.26, 505.0, 547.39, 545.25, 477.3, 433.6, 438.2, 374.7, 389.0, 440.01]
print('tesla:',prices)

def max_profit(prices):
    out = prices[1] - prices[0]
    min_price = prices[0]
    re = ()
    for i,e in enumerate(prices[1:]):
        if e - min_price > out:
            out = e - min_price
            sell = e
        if e < min_price:
            min_price = e
    return round(out,2),sell-round(out,2),sell

print('bench:',max_profit(prices))



import xlrd
# 打开 xls 文件道
book = xlrd.open_workbook(r'D:/python/excel/tesla_price.xls')
# 获取第 1 个表单
sh = book.sheet_by_index(0)
buy_in = prices.index(max_profit(prices)[1])
sell_out = prices.index(max_profit(prices)[2])

colum = 1 #开盘价
#excel第1行对应prices序列最后一个元素price[-1]
b = max_profit(prices)[1]
s = max_profit(prices)[2]
print("第"+f'{days-buy_in}'+"行第"+f'{colum}'+"列:", b)
print("第"+f'{days-sell_out}'+"行第"+f'{colum}'+"列:",s)

import datetime
def convert(dates):#定义转化日期戳的函数,dates为日期戳
    delta=datetime.timedelta(days=dates)
    today=datetime.datetime.strptime('1899-12-30','%Y-%m-%d')+delta
    #将1899-12-30转化为可以计算的时间格式并加上要转化的日期戳
    return datetime.datetime.strftime(today,'%Y-%m-%d')
    #制定输出日期的格式
colum_date = 0
buy_in_time = convert(sh.cell_value(days-buy_in, colum_date))
sell_out_time = convert(sh.cell_value(days-sell_out, colum_date))
print("买入时机:", buy_in_time,b)
print("卖出时机:", sell_out_time,s)

def max_profit(prices):
    save, best = float('inf'), float('-inf')
    for p in prices:
        best = max(best, p-save)
        save = min(save, p)
    return best
'''
The function time.clock() has been removed, after having been deprecated since Python 3.3: 
use time.perf_counter() or time.process_time() instead, depending on your requirements, 
to have well-defined behavior. (Contributed by Matthias Bussonnier in bpo-36895.)

import xlrd

#路径前加 r，读取的文件路径
file_path = r'D:/python/excel/tesla_price.xls'

#文件路径的中文转码
#file_path = file_path.decode('utf-8')

#获取数据
data = xlrd.open_workbook(file_path)

#获取sheet
table = data.sheet_by_name('Sheet1')

#获取总行数
nrows = table.nrows
#获取总列数
ncols = table.ncols


#获取一行的数值，例如第5行
rowvalue = table.row_values(5)

#获取一列的数值，例如第6列
col_values = table.col_values(6)

#获取一个单元格的数值，例如第5行第6列
cell_value = table.cell(5,6).value
print(rowvalue)
print(col_values)
def read_excel(): #文件位置
    ExcelFile=xlrd.open_workbook(r'D:\python\excel\tesla_price')
    print (ExcelFile.sheet_names())
#获取目标EXCEL文件sheet名
#------------------------------------
#若有多个sheet，则需要指定读取目标sheet例如读取sheet2
#sheet2_name=ExcelFile.sheet_names()[1]

#获取sheet内容【1.根据sheet索引2.根据sheet名称】
#sheet=ExcelFile.sheet_by_index(1)

    sheet=ExcelFile.sheet_by_name('TestCase002')
    #打印sheet的名称，行数，列数
    print (sheet.name,sheet.nrows,sheet.ncols)

#获取整行或者整列的值

    rows=sheet.row_values(2)#第三行内容
    cols=sheet.col_values(1)#第二列内容
    print (cols,rows)

#获取单元格内容

    print (sheet.cell(1,0).value.encode('utf-8'))
    print (sheet.cell_value(1,0).encode('utf-8'))
    print (sheet.row(1)[0].value.encode('utf-8'))

#打印单元格内容格式

    print (sheet.cell(1,0).ctype)
file_path = 'D/python/excel/tesla price.xlsx'

df = pd.read_excel(r'data.xlsx',sheetname=0)
print(df.head())


workbook = openpyxl.load_workbook('D:\\python\\excel\\tesla price.xlsx')
worksheet = workbook.get_sheet_by_name('Sheet1')
row3=[item.value for item in list(worksheet.rows)[2]]
print('第3行值',row3)

col3=[item.value for item in list(worksheet.columns)[2]]
print('第3行值',col3)

cell_2_3=worksheet.cell(row=2,column=3).value
print('第2行第3列值',cell_2_3)

max_row=worksheet.max_row
print('最大行',max_row)


'''